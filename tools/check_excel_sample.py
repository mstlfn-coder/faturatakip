from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.utils.cell import range_boundaries, get_column_letter


@dataclass(frozen=True)
class SheetSummary:
    name: str
    print_area: str | None
    min_row: int | None
    max_row: int | None
    min_col: int | None
    max_col: int | None
    header_hits: dict[str, list[tuple[str, str]]]
    tail_text_rows: list[tuple[int, list[str]]]


KEYS = [
    "Kurum",
    "Dönem",
    "Rapor",
    "Rapor Tarihi",
    "Filtre",
    "Hazırlayan",
    "Kontrol Eden",
    "Toplam",
    "Öden",
    "Kalan",
    "Gecik",
    "PDF",
    "Sayfa",
]


def _iter_text_cells(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> Iterable[tuple[int, int, str]]:
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                s = v.strip()
                if s:
                    yield r, c, s


def summarize_sheet(ws) -> SheetSummary:
    pa = ws.print_area
    if not pa:
        return SheetSummary(ws.title, None, None, None, None, None, {}, [])

    area = str(pa).split("!")[-1].replace("$", "")
    min_col, min_row, max_col, max_row = range_boundaries(area)

    hits: dict[str, list[tuple[str, str]]] = {}
    for r, c, s in _iter_text_cells(ws, min_row, max_row, min_col, max_col):
        low = s.lower()
        for k in KEYS:
            if k.lower() in low:
                addr = f"{get_column_letter(c)}{r}"
                hits.setdefault(k, []).append((addr, s[:160]))
                break

    tail_rows: list[tuple[int, list[str]]] = []
    for r in range(max(min_row, max_row - 12), max_row + 1):
        row_texts = [s for (_rr, _cc, s) in _iter_text_cells(ws, r, r, min_col, max_col)]
        if row_texts:
            tail_rows.append((r, row_texts[:10]))

    return SheetSummary(ws.title, str(pa), min_row, max_row, min_col, max_col, hits, tail_rows)


def main() -> None:
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path", type=str)
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    print("Workbook:", xlsx_path)
    print("Sheets:", wb.sheetnames)

    for name in wb.sheetnames:
        ws = wb[name]
        summary = summarize_sheet(ws)
        print("\n==", summary.name, "==")
        print("print_area:", summary.print_area)
        if summary.print_area:
            print(
                "bounds:",
                f"rows {summary.min_row}..{summary.max_row}, cols {get_column_letter(summary.min_col)}..{get_column_letter(summary.max_col)}",
            )
        for k in KEYS:
            if k in summary.header_hits:
                print("hit", k, "->", summary.header_hits[k][:5])
        if summary.tail_text_rows:
            print("tail text rows:")
            for r, texts in summary.tail_text_rows:
                print(" ", r, texts)


if __name__ == "__main__":
    main()

